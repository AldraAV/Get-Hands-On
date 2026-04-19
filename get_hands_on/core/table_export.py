"""
table_export.py — Extracción de tablas PDF a CSV y XLSX.
Usa el motor find_tables() de PyMuPDF (>= 1.23).
"""
import csv
import fitz
from pathlib import Path


def extract_tables_to_csv(pdf_path: str, output_dir: str) -> list[str]:
    """
    Extrae todas las tablas de un PDF y las guarda como archivos CSV individuales.
    Retorna la lista de rutas de archivos creados.
    """
    doc = fitz.open(pdf_path)
    stem = Path(pdf_path).stem
    out_dir = Path(output_dir)
    out_dir.mkdir(parents=True, exist_ok=True)

    created = []
    table_index = 1

    for page_num, page in enumerate(doc, start=1):
        try:
            tabs = page.find_tables()
        except AttributeError:
            continue

        for tab in tabs.tables:
            data = _extract_data(tab)
            if not data:
                continue

            out_path = out_dir / f"{stem}_tabla_{table_index}_pag{page_num}.csv"
            with open(out_path, "w", newline="", encoding="utf-8-sig") as f:
                writer = csv.writer(f)
                writer.writerows(data)
            created.append(str(out_path))
            table_index += 1

    doc.close()
    return created


def extract_tables_to_xlsx(pdf_path: str, output_path: str) -> int:
    """
    Extrae todas las tablas del PDF en un único archivo XLSX.
    Cada tabla va en una hoja separada (Tabla_1, Tabla_2...).
    Retorna el número de tablas exportadas.
    """
    try:
        import openpyxl
    except ImportError:
        raise ImportError("openpyxl no está instalado. Ejecuta: pip install openpyxl")

    doc = fitz.open(pdf_path)
    wb = openpyxl.Workbook()
    wb.remove(wb.active)  # remover hoja vacía por defecto

    table_index = 1

    for page_num, page in enumerate(doc, start=1):
        try:
            tabs = page.find_tables()
        except AttributeError:
            continue

        for tab in tabs.tables:
            data = _extract_data(tab)
            if not data:
                continue

            ws = wb.create_sheet(title=f"Tabla_{table_index}_P{page_num}")

            for row in data:
                ws.append(row)

            # Aplicar estilos básicos al encabezado (primera fila)
            from openpyxl.styles import Font, PatternFill, Alignment
            header_fill = PatternFill("solid", fgColor="2D2D2D")
            for cell in ws[1]:
                cell.font = Font(bold=True, color="FF6B00")
                cell.fill = header_fill
                cell.alignment = Alignment(horizontal="center")

            table_index += 1

    doc.close()

    if table_index == 1:
        return 0  # No se encontraron tablas

    wb.save(output_path)
    return table_index - 1


def _extract_data(tab) -> list[list[str]]:
    """Extrae contenido de un objeto Table de PyMuPDF como lista de listas."""
    try:
        raw = tab.extract()
    except Exception:
        return []
    if not raw:
        return []
    return [
        [str(cell).replace("\n", " ").strip() if cell is not None else "" for cell in row]
        for row in raw
    ]
